import io
import pandas as pd
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import func

# PDF generation imports
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# Excel generation imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from clv_platform.database.models import Customer, CustomerClvPrediction, CustomerSegment, Transaction

def generate_pdf_report(db: Session) -> io.BytesIO:
    """
    Generates a professional, print-ready PDF executive report summarizing CLV platform data.
    """
    # 1. Gather Summary Data
    total_customers = db.query(Customer).count()
    
    clv_stats = db.query(
        func.avg(CustomerClvPrediction.predicted_clv_6months),
        func.sum(CustomerClvPrediction.predicted_clv_6months),
        func.avg(CustomerClvPrediction.churn_risk_score)
    ).first()
    
    avg_clv = float(clv_stats[0]) if clv_stats[0] is not None else 0.0
    total_clv = float(clv_stats[1]) if clv_stats[1] is not None else 0.0
    avg_churn = float(clv_stats[2]) if clv_stats[2] is not None else 0.0
    
    # Tier distribution counts
    tier_counts = db.query(
        CustomerClvPrediction.recommendation_tier,
        func.count(CustomerClvPrediction.id)
    ).group_by(CustomerClvPrediction.recommendation_tier).all()
    
    tier_map = {t[0]: t[1] for t in tier_counts}
    
    # Top 10 High CLV Customers
    top_10 = db.query(
        CustomerClvPrediction.customer_id,
        CustomerClvPrediction.predicted_clv_6months,
        CustomerClvPrediction.churn_risk_score,
        CustomerClvPrediction.recommendation_tier
    ).order_by(CustomerClvPrediction.predicted_clv_6months.desc()).limit(10).all()

    # 2. Document Setup
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=letter,
        rightMargin=54,
        leftMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    # Styling
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.HexColor('#1A365D'),
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#4A5568'),
        spaceAfter=30
    )
    
    section_title = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#2C5282'),
        spaceBefore=15,
        spaceAfter=10
    )
    
    body_style = ParagraphStyle(
        'DocBody',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#2D3748'),
        spaceAfter=8
    )

    story = []
    
    # Header block
    story.append(Paragraph("Executive CLV & Customer Lifetime Analytics Report", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Platform Version: 2.0.0", subtitle_style))
    story.append(Spacer(1, 10))
    
    # Executive Summary Section
    story.append(Paragraph("1. Executive Summary", section_title))
    summary_text = (
        f"This business intelligence report summarizes the predicted Customer Lifetime Value (CLV) "
        f"and customer engagement metrics computed across the active database. Currently, "
        f"<b>{total_customers:,}</b> active customer profiles are tracked in the system. "
        f"The average predicted 6-month revenue value per customer is <b>£{avg_clv:.2f}</b>, "
        f"projecting a total customer pipeline value of <b>£{total_clv:,.2f}</b> over the next 180 days. "
        f"The average estimated portfolio churn risk is <b>{avg_churn * 100:.1f}%</b>."
    )
    story.append(Paragraph(summary_text, body_style))
    story.append(Spacer(1, 15))
    
    # KPI Grid Table
    kpi_data = [
        [Paragraph("<b>KPI Metric</b>", body_style), Paragraph("<b>Value</b>", body_style)],
        ["Total Customer Database", f"{total_customers:,}"],
        ["Projected 6-Month Revenue Pool", f"£{total_clv:,.2f}"],
        ["Mean Customer Value (6m)", f"£{avg_clv:.2f}"],
        ["Average Customer Churn Probability", f"{avg_churn * 100:.1f}%"]
    ]
    kpi_table = Table(kpi_data, colWidths=[250, 150])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#EDF2F7')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#2D3748')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
        ('BOTTOMPADDING', (0,1), (-1,-1), 5),
        ('TOPPADDING', (0,1), (-1,-1), 5),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 20))
    
    # Segment Distribution Section
    story.append(Paragraph("2. Customer Segment Tiers Distribution", section_title))
    segment_text = (
        "Customers are automatically mapped into four strategic tiers based on their predicted CLV: "
        "Platinum, Gold, Silver, and Bronze. Each tier maps to distinct marketing strategies."
    )
    story.append(Paragraph(segment_text, body_style))
    story.append(Spacer(1, 10))
    
    # Segment Table
    seg_table_data = [[
        Paragraph("<b>CLV Tier</b>", body_style), 
        Paragraph("<b>Active Customers</b>", body_style),
        Paragraph("<b>Percentage Share</b>", body_style)
    ]]
    for tier in ["Platinum", "Gold", "Silver", "Bronze"]:
        cnt = tier_map.get(tier, 0)
        pct = (cnt / total_customers * 100) if total_customers > 0 else 0
        seg_table_data.append([tier, f"{cnt:,}", f"{pct:.1f}%"])
        
    seg_table = Table(seg_table_data, colWidths=[150, 120, 130])
    seg_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#EDF2F7')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
    ]))
    story.append(seg_table)
    
    story.append(PageBreak())
    
    # Top Customers Section
    story.append(Paragraph("3. Top 10 High-Value Customers (Target List)", section_title))
    top_cust_text = (
        "The table below list the top 10 customers based on their 6-month predicted CLV. "
        "These customers represent the highest-priority targets for VIP retention and loyalty marketing."
    )
    story.append(Paragraph(top_cust_text, body_style))
    story.append(Spacer(1, 10))
    
    top_table_data = [[
        Paragraph("<b>Customer ID</b>", body_style),
        Paragraph("<b>Predicted CLV (6m)</b>", body_style),
        Paragraph("<b>Churn Risk</b>", body_style),
        Paragraph("<b>Strategic Tier</b>", body_style)
    ]]
    for item in top_10:
        top_table_data.append([
            item[0],
            f"£{float(item[1]):,.2f}",
            f"{float(item[2]) * 100:.1f}%",
            item[3]
        ])
        
    top_table = Table(top_table_data, colWidths=[100, 110, 90, 100])
    top_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#EDF2F7')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E0')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#FEFCBF')), # Highlight top 1
    ]))
    story.append(top_table)
    story.append(Spacer(1, 20))
    
    # Closing signature/disclaimer
    story.append(Paragraph("4. Recommended Retention Strategies", section_title))
    rec_text = (
        "• <b>Platinum Tiers</b>: Immediately enroll in VIP concierge campaign. High lifetime value warrants premium personalized gifts and direct sales calls.<br/>"
        "• <b>Gold Tiers</b>: Nurture with free shipments, loyalty reward invites, and early brand releases.<br/>"
        "• <b>Silver Tiers</b>: Drive purchase frequency through bundle values and recommended complement items.<br/>"
        "• <b>Bronze Tiers</b>: Standard welcome emails, seasonal discount offerings, and re-engagement campaigns."
    )
    story.append(Paragraph(rec_text, body_style))
    
    # Build
    doc.build(story)
    pdf_buffer.seek(0)
    return pdf_buffer

def generate_excel_report(db: Session) -> io.BytesIO:
    """
    Generates a multi-sheet, beautifully formatted Excel workbook with deep CLV and segments details.
    """
    wb = Workbook()
    
    # Fonts and Styling
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    accent_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # Sheet 1: Dashboard Overview
    ws_dash = wb.active
    ws_dash.title = "Overview Metrics"
    ws_dash.views.sheetView[0].showGridLines = True
    
    total_customers = db.query(Customer).count()
    clv_stats = db.query(
        func.avg(CustomerClvPrediction.predicted_clv_6months),
        func.sum(CustomerClvPrediction.predicted_clv_6months),
        func.avg(CustomerClvPrediction.churn_risk_score)
    ).first()
    
    avg_clv = float(clv_stats[0]) if clv_stats[0] is not None else 0.0
    total_clv = float(clv_stats[1]) if clv_stats[1] is not None else 0.0
    avg_churn = float(clv_stats[2]) if clv_stats[2] is not None else 0.0
    
    ws_dash.merge_cells("A1:C1")
    ws_dash["A1"] = "CLV PLATFORM EXECUTIVE SUMMARY"
    ws_dash["A1"].font = title_font
    ws_dash["A1"].fill = header_fill
    ws_dash["A1"].alignment = Alignment(horizontal="center")
    
    ws_dash["A3"] = "Metric Key"
    ws_dash["A3"].font = header_font
    ws_dash["A3"].fill = accent_fill
    ws_dash["B3"] = "Value"
    ws_dash["B3"].font = header_font
    ws_dash["B3"].fill = accent_fill
    
    metrics_list = [
        ("Total Database Customers", total_customers),
        ("Average Projected 6M CLV", avg_clv),
        ("Total Projected Revenue Pool (6M)", total_clv),
        ("Average Customer Churn Risk", avg_churn)
    ]
    
    for idx, (m_key, m_val) in enumerate(metrics_list, start=4):
        ws_dash[f"A{idx}"] = m_key
        ws_dash[f"B{idx}"] = m_val
        ws_dash[f"A{idx}"].border = thin_border
        ws_dash[f"B{idx}"].border = thin_border
        
        # Formats
        if "CLV" in m_key or "Revenue" in m_key:
            ws_dash[f"B{idx}"].number_format = '"£"#,##0.00'
        elif "Risk" in m_key:
            ws_dash[f"B{idx}"].number_format = "0.0%"
        else:
            ws_dash[f"B{idx}"].number_format = "#,##0"

    # Sheet 2: Top Customers Sheet
    ws_top = wb.create_sheet(title="Top 500 Spenders")
    ws_top.views.sheetView[0].showGridLines = True
    
    top_500_query = (
        db.query(
            CustomerClvPrediction.customer_id,
            CustomerClvPrediction.predicted_clv_6months,
            CustomerClvPrediction.churn_risk_score,
            CustomerClvPrediction.churn_risk_tier,
            CustomerClvPrediction.expected_purchases_6m,
            CustomerClvPrediction.recommendation_tier
        )
        .order_by(CustomerClvPrediction.predicted_clv_6months.desc())
        .limit(500)
    )
    
    df_top = pd.read_sql(top_500_query.statement, db.bind)
    
    if not df_top.empty:
        df_top.columns = ["Customer ID", "Predicted CLV (6m)", "Churn Risk Score", "Churn Risk Tier", "Expected Purchases (6m)", "CLV Tier"]
        for r in dataframe_to_rows(df_top, index=False, header=True):
            ws_top.append(r)
            
        # Format Top headers
        for cell in ws_top[1]:
            cell.font = header_font
            cell.fill = accent_fill
            cell.border = thin_border
            
        for row in range(2, ws_top.max_row + 1):
            ws_top[f"B{row}"].number_format = '"£"#,##0.00'
            ws_top[f"C{row}"].number_format = '0.0%'
            ws_top[f"E{row}"].number_format = '#,##0.0'
            for col in ["A", "B", "C", "D", "E", "F"]:
                ws_top[f"{col}{row}"].border = thin_border

    # Sheet 3: At Risk Customers Sheet
    ws_risk = wb.create_sheet(title="At-Risk High-Value")
    ws_risk.views.sheetView[0].showGridLines = True
    
    at_risk_query = (
        db.query(
            CustomerClvPrediction.customer_id,
            CustomerClvPrediction.predicted_clv_6months,
            CustomerClvPrediction.churn_risk_score,
            CustomerClvPrediction.recommendation_tier
        )
        .filter(CustomerClvPrediction.churn_risk_tier == "High")
        .order_by(CustomerClvPrediction.predicted_clv_6months.desc())
        .limit(500)
    )
    
    df_risk = pd.read_sql(at_risk_query.statement, db.bind)
    
    if not df_risk.empty:
        df_risk.columns = ["Customer ID", "Predicted CLV (6m)", "Churn Risk Score", "CLV Tier"]
        for r in dataframe_to_rows(df_risk, index=False, header=True):
            ws_risk.append(r)
            
        for cell in ws_risk[1]:
            cell.font = header_font
            cell.fill = accent_fill
            cell.border = thin_border
            
        for row in range(2, ws_risk.max_row + 1):
            ws_risk[f"B{row}"].number_format = '"£"#,##0.00'
            ws_risk[f"C{row}"].number_format = '0.0%'
            for col in ["A", "B", "C", "D"]:
                ws_risk[f"{col}{row}"].border = thin_border

    # Auto-adjust column widths for all sheets
    from openpyxl.utils import get_column_letter
    for ws in [ws_dash, ws_top, ws_risk]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column) if isinstance(col[0].column, int) else col[0].column
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Save to memory stream
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer
